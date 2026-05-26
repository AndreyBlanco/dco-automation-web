"""Read/write the plan catalog and verification log in Excel (openpyxl)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import get_excel_path

PLAN_SHEET = "plan_catalog"
LOG_SHEET = "verification_log"

PLAN_HEADERS = ["carrier", "plan_name", "portal_url"]
LOG_HEADERS = [
    "completed_at",
    "job_id",
    "patient_id",
    "member_id",
    "carrier",
    "plan_name",
    "status",
    "source",
    "first_name",
    "last_name",
]

SEED_PLANS: list[tuple[str, str, str]] = [
    ("Aetna", "Aetna Dental PPO", "https://www.aetna.com"),
    ("SmileCare", "SmileCare PPO", "https://example.com/smilecare"),
    ("FamilyDental", "FamilyDental Select", "https://example.com/familydental"),
    ("Delta Dental", "Delta Dental PPO", "https://www.deltadental.com"),
]


@dataclass(frozen=True)
class PlanRow:
    carrier: str
    plan_name: str
    portal_url: str | None = None


def _normalize(value: str) -> str:
    return " ".join(value.strip().lower().split())


def ensure_workbook(path: Path | None = None) -> Path:
    excel_path = path or get_excel_path()
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        return excel_path

    wb = Workbook()
    plan_ws = wb.active
    plan_ws.title = PLAN_SHEET
    plan_ws.append(PLAN_HEADERS)
    for carrier, plan_name, portal_url in SEED_PLANS:
        plan_ws.append([carrier, plan_name, portal_url])

    log_ws = wb.create_sheet(LOG_SHEET)
    log_ws.append(LOG_HEADERS)
    wb.save(excel_path)
    return excel_path


def _plan_sheet(wb: Workbook) -> Worksheet:
    if PLAN_SHEET in wb.sheetnames:
        return wb[PLAN_SHEET]
    ws = wb.create_sheet(PLAN_SHEET)
    ws.append(PLAN_HEADERS)
    return ws


def _log_sheet(wb: Workbook) -> Worksheet:
    if LOG_SHEET in wb.sheetnames:
        return wb[LOG_SHEET]
    ws = wb.create_sheet(LOG_SHEET)
    ws.append(LOG_HEADERS)
    return ws


def list_plans(path: Path | None = None) -> list[PlanRow]:
    excel_path = ensure_workbook(path)
    wb = load_workbook(excel_path, read_only=True)
    ws = _plan_sheet(wb)
    rows: list[PlanRow] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        carrier = str(row[0]).strip()
        plan_name = str(row[1]).strip() if row[1] else carrier
        portal_url = str(row[2]).strip() if len(row) > 2 and row[2] else None
        rows.append(PlanRow(carrier=carrier, plan_name=plan_name, portal_url=portal_url))
    wb.close()
    return rows


def lookup_plan(carrier: str, plan_name: str, path: Path | None = None) -> PlanRow | None:
    nc, np = _normalize(carrier), _normalize(plan_name)
    for row in list_plans(path):
        if _normalize(row.carrier) == nc and _normalize(row.plan_name) == np:
            return row
    return None


def add_plan(
    carrier: str,
    plan_name: str,
    portal_url: str | None = None,
    path: Path | None = None,
) -> PlanRow:
    existing = lookup_plan(carrier, plan_name, path)
    if existing:
        return existing

    excel_path = ensure_workbook(path)
    wb = load_workbook(excel_path)
    ws = _plan_sheet(wb)
    url = portal_url or f"https://portal.{_normalize(carrier).replace(' ', '')}.example"
    ws.append([carrier.strip(), plan_name.strip(), url])
    wb.save(excel_path)
    wb.close()
    return PlanRow(carrier=carrier.strip(), plan_name=plan_name.strip(), portal_url=url)


def append_verification_log(
    *,
    job_id: str,
    patient_id: str,
    member_id: str,
    carrier: str,
    plan_name: str,
    status: str,
    source: str,
    first_name: str,
    last_name: str,
    path: Path | None = None,
) -> None:
    excel_path = ensure_workbook(path)
    wb = load_workbook(excel_path)
    ws = _log_sheet(wb)
    completed_at = datetime.now(timezone.utc).isoformat()
    ws.append(
        [
            completed_at,
            job_id,
            patient_id,
            member_id,
            carrier,
            plan_name,
            status,
            source,
            first_name,
            last_name,
        ]
    )
    wb.save(excel_path)
    wb.close()
